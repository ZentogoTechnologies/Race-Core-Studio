"""Saca el Excel de pilotos, una pestaña por categoría.

No es un informe sino una plantilla de captura: la idea es repartir el
archivo por categorías para que cada quien complete lo suyo —equipo,
nacionalidad, foto— y después volver a cargarlo en la base.

    python exportar_pilotos.py
    python exportar_pilotos.py --salida "C:/ruta/Pilotos.xlsx"

Lee por el API y no por Mongo directamente, para que salga exactamente lo
que ve la aplicación y no una lectura paralela que pueda desviarse.
"""

import argparse
import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

API = "http://127.0.0.1:8080/api/v1"

CARPETA_FOTOS = Path(__file__).resolve().parent / "src" / "public" / "pilotos"

EXTENSIONES = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}

# Las columnas de cada pestaña. El ID va primero y en gris: es la única
# forma de saber si una fila es un piloto que ya existe o uno nuevo.
COLUMNAS = [
    ("ID", 8),
    ("Nombre", 20),
    ("Apellido", 24),
    ("Nacionalidad", 18),
    ("Equipo / Marca", 26),
    ("Foto (archivo)", 26),
    ("También corre en", 26),
    ("Notas", 34),
]

ROJO = "C00000"
GRIS = "3F3F3F"
CREMA = "FFF2CC"

FINA = Side(style="thin", color="D9D9D9")
BORDE = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)

INSTRUCCIONES = [
    ("Cómo llenar este archivo", None),
    ("", None),
    ("Hay una pestaña por categoría. Cada quien completa la suya.", None),
    ("", None),
    ("ID",
     "No lo toques. Es como el sistema reconoce a un piloto que ya existe. "
     "Si agregas un piloto nuevo, deja la celda vacía: el número se asigna solo al cargar."),
    ("Nombre y Apellido",
     "En columnas separadas, y el apellido completo. Si es compuesto, escríbelo entero "
     "en la columna de apellido: DE GRACIA, JEAN FRANCOIS. Es lo que no se puede adivinar después."),
    ("Nacionalidad",
     "El país, escrito igual en todas las filas: Panama, Colombia, Venezuela. "
     "Sale en la ficha del piloto."),
    ("Equipo / Marca",
     "El equipo al que corre. Hoy faltan en más de la mitad de los pilotos."),
    ("Foto (archivo)",
     "El nombre del archivo de la foto, con su extensión: 'Alan De Gracia.jpg'. "
     "Las fotos van aparte, todas en una carpeta. Nómbralas con el nombre y apellido "
     "del piloto tal como los escribiste aquí."),
    ("También corre en",
     "Informativo. Dice en qué otras categorías está ese piloto ahora mismo."),
    ("Notas", "Para lo que haga falta aclarar. No se carga a la base."),
    ("", None),
    ("Un piloto que corre en dos categorías", None),
    ("",
     "Va en las dos pestañas, con el nombre y el apellido escritos exactamente igual "
     "en las dos. Así se carga como una sola persona en ambas categorías, y no como "
     "dos pilotos distintos. Si ya tiene ID, repite el mismo ID en las dos filas."),
    ("", None),
    ("Lo que no hay que cambiar", None),
    ("",
     "El nombre de las pestañas ni la fila de títulos. Si se renombran, la carga no "
     "encuentra las columnas. Agregar filas al final sí, sin problema."),
]


def fotos_en_disco():
    """Qué foto tiene cada piloto, mirando la carpeta.

    El campo `photo` de la base casi nunca está puesto: las fotos que hay
    se llaman como el id del piloto y el backend las encuentra buscando.
    Si la columna se llenara solo con ese campo, el Excel diría que no hay
    foto en pilotos que sí la tienen, y alguien la buscaría dos veces.
    """
    encontradas = {}

    if not CARPETA_FOTOS.is_dir():
        return encontradas

    for archivo in CARPETA_FOTOS.rglob("*"):
        if not archivo.is_file() or archivo.suffix.lower() not in EXTENSIONES:
            continue
        if not archivo.stem.isdigit():
            continue

        pid = int(archivo.stem)
        # El primero que aparece manda; no debería haber dos para el mismo
        # piloto y si los hay, es algo que revisar aparte.
        encontradas.setdefault(pid, archivo.name)

    return encontradas


def pedir(ruta, cuerpo=None, token=None):
    datos = json.dumps(cuerpo, ensure_ascii=False).encode("utf-8") if cuerpo is not None else None
    peticion = urllib.request.Request(API + ruta, data=datos)
    if datos is not None:
        peticion.add_header("Content-Type", "application/json")
    if token:
        peticion.add_header("Authorization", "Bearer " + token)
    with urllib.request.urlopen(peticion, timeout=60) as resp:
        return json.loads(resp.read().decode())


def hoja_instrucciones(wb):
    ws = wb.create_sheet("INSTRUCCIONES", 0)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96

    fila = 1
    for titulo, texto in INSTRUCCIONES:
        a = ws.cell(row=fila, column=1, value=titulo)
        if texto is None and titulo:
            # Encabezado de sección
            a.font = Font(bold=True, size=13, color=ROJO)
        else:
            a.font = Font(bold=True)
            b = ws.cell(row=fila, column=2, value=texto)
            b.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[fila].height = 30
        fila += 1

    return ws


def hoja_categoria(wb, nombre, pilotos, nombres_categorias, fotos):
    # Excel no admite estos caracteres en el nombre de una pestaña, y el
    # tope son 31 caracteres.
    limpio = nombre
    for c in "[]:*?/\\":
        limpio = limpio.replace(c, "-")
    ws = wb.create_sheet(limpio[:31])

    ws.append([t for t, _ in COLUMNAS])

    cabecera = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor=ROJO)
    for i in range(1, len(COLUMNAS) + 1):
        celda = ws.cell(row=1, column=i)
        celda.font = cabecera
        celda.fill = fondo
        celda.alignment = Alignment(horizontal="center", vertical="center")

    gris = Font(color=GRIS)
    for p in pilotos:
        otras = [nombres_categorias.get(c, f"#{c}")
                 for c in (p.get("categories") or [])
                 if nombres_categorias.get(c) != nombre]

        ws.append([
            p.get("pilot_id"),
            p.get("name", ""),
            p.get("last_name", ""),
            p.get("nationality") or "",
            p.get("team_brand") or "",
            (Path(p["photo"]).name if p.get("photo")
             else fotos.get(p.get("pilot_id"), "")),
            ", ".join(otras),
            "",
        ])

        f = ws.max_row
        ws.cell(row=f, column=1).font = gris          # el ID, apagado
        ws.cell(row=f, column=7).font = gris          # informativo

        # Lo que falta se pinta, para que se vea de un vistazo qué hay
        # que completar sin ir celda por celda.
        for col in (4, 5, 6):
            celda = ws.cell(row=f, column=col)
            if not celda.value:
                celda.fill = PatternFill("solid", fgColor=CREMA)

        for col in range(1, len(COLUMNAS) + 1):
            ws.cell(row=f, column=col).border = BORDE

    for i, (_, ancho) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{max(ws.max_row, 1)}"

    return ws


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--salida", default="Pilotos Race Core Studio.xlsx")
    parser.add_argument("--usuario", default="zentogotech")
    parser.add_argument("--password", default="zentogotech")
    args = parser.parse_args()

    try:
        token = pedir("/auth/login",
                      {"username": args.usuario, "password": args.password})["access_token"]
    except urllib.error.URLError as e:
        print(f"No se pudo hablar con el backend en {API}: {e}")
        print("Arráncalo antes de exportar.")
        return 1

    categorias = pedir("/categories/", token=token)["items"]
    nombres = {c["category_id"]: c["category_name"] for c in categorias}

    pilotos = pedir("/pilots/?sort_by=last_name", token=token)["items"]

    fotos = fotos_en_disco()

    wb = Workbook()
    wb.remove(wb.active)

    hoja_instrucciones(wb)

    total_filas = 0
    for c in sorted(categorias, key=lambda x: x["category_name"]):
        cid = c["category_id"]
        de_esta = [p for p in pilotos if cid in (p.get("categories") or [])]
        hoja_categoria(wb, c["category_name"], de_esta, nombres, fotos)
        total_filas += len(de_esta)
        print(f"  {c['category_name']:<18} {len(de_esta):>3} pilotos")

    # Los que no están en ninguna categoría no pueden perderse en la
    # exportación: son justo los que hay que revisar.
    sueltos = [p for p in pilotos if not (p.get("categories") or [])]
    if sueltos:
        hoja_categoria(wb, "Sin categoria", sueltos, nombres, fotos)
        print(f"  {'Sin categoría':<18} {len(sueltos):>3} pilotos")

    destino = Path(args.salida)
    wb.save(destino)

    print(f"\n  {destino}")
    con_foto = sum(1 for p in pilotos
                   if p.get("photo") or fotos.get(p.get("pilot_id")))

    print(f"  {len(pilotos)} pilotos, {total_filas + len(sueltos)} filas "
          f"(un piloto de dos categorías sale en las dos)")
    print(f"  {con_foto} con foto, {len(pilotos) - con_foto} sin ella")
    return 0


if __name__ == "__main__":
    sys.exit(main())
